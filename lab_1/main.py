from datetime import datetime
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import openpyxl

workbook = openpyxl.load_workbook("/home/user/Рабочий стол/students/607-12/VerkhoglyadovVN/py_labs/spisok.xlsx")
worksheet = workbook.active
for i in range (2,10):
    student_name = worksheet.cell(row=i, column=1).value + ' ' + worksheet.cell(row=i, column=2).value
    student_nomination = worksheet.cell(row=i, column=3).value
    sign_date = str(worksheet.cell(row=i, column = 4).value.date())
    signed_by = worksheet.cell(row=i, column =5).value

    photo = Image.open('/home/user/Рабочий стол/students/607-12/VerkhoglyadovVN/py_labs/test.jpg')
    drawing = ImageDraw.Draw(photo)
    font = ImageFont.truetype('/home/user/Рабочий стол/students/607-12/VerkhoglyadovVN/py_labs/Шрифт.otf',40)
    for i in range(2,10):
    
        drawing.text((240,300), 'Награждается',(3,8,12),font=font)
        drawing.text((220,380), student_name,(3,8,12), font=font)
        drawing.text((260,460), 'В номинации',(3,8,12), font=font)
        drawing.text((100,540), student_nomination,(3,8,12), font=font)
        drawing.text((80,810), 'Дата:',(3,8,12), font=font)
        drawing.text((80,860), sign_date,(3,8,12), font=font)
        drawing.text((450,810), 'Подписал:',(3,8,12), font=font)
        drawing.text((450,860), signed_by,(3,8,12), font=font)

    photo.save('/home/user/Рабочий стол/students/607-12/VerkhoglyadovVN/py_labs/diplomes/'+student_name+'.jpg')